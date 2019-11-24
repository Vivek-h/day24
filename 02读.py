from openpyxl import load_workbook
wb = load_workbook('s15.xlsx',data_only=True)
#获取所有的工作表
print(wb.sheetnames)

wb1=wb['user']
#方式一：通过索引
a= wb1['A1']
print(a.value)

#方式二
print(wb1.cell(row=1,column=2).value)

#求最大行
print(wb1.max_row)

#求最大列
print(wb1.max_column)

#求每一行的值
print(wb1.rows)

#求每一列的值
print(wb1.columns)