from openpyxl import Workbook
wb=Workbook()

#方式一：默认在最后
# wb1=wb.create_sheet('index')

#方式二：根据索引的位置来添加工作表
wb1=wb.create_sheet('index',0)
#方式一：添加内容，用单元格的索引来添加
# wb1['D3'] = '停车坐爱枫林晚，霜叶红于二月花'

#方式二：根据单元格的位置来添加
# wb1.cell(row=3,column=5,value='先帝创业为伴而中道崩殂，今天下三分益州疲敝')

#函数
# wb1['A1']=4
# wb1['A2']=6
# wb1['A3']='=sum(A1:A2)'

#添加行
l=['姓名','性别','年龄','爱好','住址','电话']
wb1.append(l)

wb1.title='user'

wb.save('s15.xlsx')